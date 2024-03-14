import openpyxl
from openpyxl.utils import get_column_letter

class ExcelWriter:
    def __init__(self, filename):
        """
        Initializes the ExcelWriter with a filename.
        Tries to load the workbook if it exists, otherwise creates a new workbook.
        """
        self.filename = filename
        try:
            self.workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()

        self.sheet = self.workbook.active

    def insert_pairs(self, pairs):
        """
        Takes a list of string pairs (artist, title) and inserts them into the Excel sheet.
        Each pair goes into a new row, starting from the first row.
        """
        row = 1
        for artist, title in pairs:
            artist_cell = 'A' + str(row)
            title_cell = 'B' + str(row)
            self.sheet[artist_cell] = artist
            self.sheet[title_cell] = title
            row += 1  # Move to the next row

        self.workbook.save(self.filename)

    def bundle_as_pairs(self, input_list):
        """
        Bundles a list of strings into pairs. Assumes the list length is even.
        If the list length is odd, the last element is ignored.
        """
        return [(input_list[i], input_list[i + 1]) for i in range(0, len(input_list) - 1, 2)]

    def insert_text(self, text):
        """
        Splits the text by newline and inserts each line into a new row in the Excel sheet.
        """
        lines = text.strip().split('\n')  # Splitting the text by newline characters
        row = 1
        for line in lines:
            self.sheet.cell(row=row, column=1, value=line)
            row += 1  # Move to the next row
        self.workbook.save(self.filename)

# Example usage:
if __name__ == "__main__":
    input_list = ['Artist1', 'Title1', 'Artist2', 'Title2', 'Artist3', 'Title3']
    excel_writer = ExcelWriter('example.xlsx')
    pairs = excel_writer.bundle_as_pairs(input_list)
    excel_writer.insert_pairs(pairs)
